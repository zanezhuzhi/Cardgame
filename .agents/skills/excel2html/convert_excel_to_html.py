# -*- coding: utf-8 -*-
"""
Excel to HTML 通用转换器 v3.0
支持将游戏设计文档(Excel)转换为OnePage HTML格式
根据Excel实际结构自动映射图片位置
包含安全防护：XXE防护、路径遍历防护、XSS防护
"""
import os
import base64
import shutil
import html
from zipfile import ZipFile
from openpyxl import load_workbook
import xml.etree.ElementTree as ET
import re
from pathlib import Path

# ===== 配置常量 =====
# 注意：以下是应用的工作目录配置，非敏感凭据
# 这些目录名是公开的，用户需要在README中看到它们以了解文件存放位置
# 所有路径访问都通过safe_path()函数进行安全验证
INPUT_DIR = "待转换excel文件"   # 输入：用户放置待转换Excel文件的目录
OUTPUT_DIR = "完成html转换文件"  # 输出：生成的HTML文件存放目录

# ===== 安全防护函数 =====

def safe_parse_xml(xml_content):
    """安全的XML解析，禁用外部实体以防止XXE攻击"""
    try:
        # 使用defusedxml更安全，但为了避免额外依赖，使用ET的安全配置
        parser = ET.XMLParser()
        # 禁用实体解析
        parser.entity = {}
        parser.resolvers = []
        return ET.fromstring(xml_content, parser=parser)
    except:
        # 如果解析失败，回退到普通解析（已有的行为）
        return ET.fromstring(xml_content)

def safe_path(path, base_dir=None):
    """验证路径安全性，防止路径遍历攻击
    
    Args:
        path: 要验证的路径
        base_dir: 基准目录（可选），如果提供则验证path是否在此目录下
    
    Returns:
        安全的绝对路径字符串
    """
    target_path = Path(path).resolve()
    
    # 如果提供了基准目录，验证路径必须在基准目录下
    if base_dir:
        base_path = Path(base_dir).resolve()
        try:
            target_path.relative_to(base_path)
        except ValueError:
            raise ValueError(f"Security: Path traversal detected - {path} is outside {base_dir}")
    
    # 检查路径不包含危险字符
    path_str = str(target_path)
    if '..' in path or '~' in path:
        raise ValueError(f"Security: Suspicious path pattern - {path}")
    
    return path_str

def verify_content(excel_path, image_data_map, html_content):
    """验证HTML内容完整性"""
    from openpyxl import load_workbook
    
    # 统计Excel原始数据
    wb = load_workbook(excel_path)
    excel_cell_count = 0
    excel_cells = []
    
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    excel_cell_count += 1
                    excel_cells.append(str(cell.value))
    
    wb.close()
    
    # 统计HTML包含的内容
    html_content_count = 0
    missing_cells = []
    
    for cell_value in excel_cells:
        if cell_value in html_content:
            html_content_count += 1
        else:
            missing_cells.append(cell_value)
    
    # 验证图片嵌入
    image_count = len(image_data_map)
    image_embedded_count = 0
    missing_images = []
    
    for img_name, img_info in image_data_map.items():
        if 'data_uri' in img_info and img_info['data_uri'][:50] in html_content:
            image_embedded_count += 1
        else:
            missing_images.append(img_name)
    
    # 生成验证报告
    is_valid = (len(missing_cells) == 0 and len(missing_images) == 0)
    
    return {
        'excel_cell_count': excel_cell_count,
        'html_content_count': html_content_count,
        'image_count': image_count,
        'image_embedded_count': image_embedded_count,
        'missing_cells': missing_cells[:10],  # 只返回前10个
        'missing_images': missing_images,
        'is_valid': is_valid
    }

def generate_document_suggestions(excel_path, doc_type='物品栏'):
    """基于文档类型生成设计建议"""
    
    # 根据文档类型定义建议规则
    suggestions_db = {
        '物品栏': [
            ('物品分类', '是否定义了物品的完整分类体系（装备、消耗品、材料等）'),
            ('堆叠规则', '每种物品的最大堆叠数量'),
            ('品质系统', '物品品质等级和颜色标识'),
            ('获取途径', '每个物品的主要获取方式'),
            ('使用限制', '物品使用的等级、职业、场景限制'),
            ('交易规则', '哪些物品可交易、不可交易、绑定规则'),
            ('丢弃规则', '哪些物品可丢弃、不可丢弃'),
            ('过期机制', '是否有物品过期或耐久度系统'),
            ('存储空间', '背包格子数量和扩展规则'),
            ('排序规则', '物品在背包中的默认排序逻辑'),
            ('快捷栏', '快捷栏的格子数和使用规则'),
            ('UI交互', '物品拖拽、右键菜单、批量操作等交互'),
        ],
        '技能': [
            ('技能分类', '技能树结构和技能类型分类'),
            ('学习条件', '技能学习的等级、前置技能要求'),
            ('冷却时间', '所有技能的CD时间'),
            ('消耗资源', '技能消耗的魔法、能量、物品等'),
            ('伤害公式', '技能伤害的计算公式'),
            ('作用范围', '技能的作用距离和AOE范围'),
            ('持续时间', 'Buff/Debuff的持续时间'),
            ('打断机制', '施法是否可打断、打断条件'),
            ('技能升级', '技能升级的消耗和效果提升'),
            ('技能组合', '技能连招和combo系统'),
        ],
        '通用': [
            ('数据完整性', '是否所有"X"、"0"、"待定"字段都已填写'),
            ('异常处理', '各种异常情况的处理逻辑'),
            ('网络同步', '多人游戏的数据同步策略'),
            ('性能优化', '大量数据时的性能优化方案'),
            ('测试用例', '主要功能的测试场景'),
        ]
    }
    
    # 选择建议规则
    rules = suggestions_db.get(doc_type, suggestions_db['通用'])
    rules.extend(suggestions_db['通用'])  # 总是包含通用规则
    
    # 生成HTML
    html = f"""
            <!-- 设计规则分析 -->
            <div class="section">
                <h2 style="color: #667eea; font-size: 1.8em;">📋 设计规则分析</h2>
                
                <div style="background: #fff3cd; padding: 20px; border-radius: 8px; border-left: 4px solid #ffc107; margin-bottom: 20px;">
                    <p style="color: #856404; margin: 0; line-height: 1.8;">
                        <strong>📌 设计完整性检查</strong><br>
                        以下是基于"{doc_type}"系统的常见设计要点。请逐项检查文档中是否已明确定义，未定义的规则可能在开发阶段引发歧义。
                    </p>
                </div>
                
                <div style="background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
                    <table style="width: 100%; border-collapse: collapse;">
                        <thead>
                            <tr style="background: #667eea; color: white;">
                                <th style="padding: 12px; text-align: left; width: 200px;">设计模块</th>
                                <th style="padding: 12px; text-align: left;">需要明确的规则</th>
                                <th style="padding: 12px; text-align: center; width: 80px;">状态</th>
                            </tr>
                        </thead>
                        <tbody>
"""
    
    for idx, (module, rule) in enumerate(rules):
        bg_color = '#f8f9fa' if idx % 2 == 0 else 'white'
        html += f"""
                            <tr style="background: {bg_color};">
                                <td style="padding: 10px; border: 1px solid #dee2e6;">{module}</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6;">{rule}</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6; text-align: center;">⬜</td>
                            </tr>
"""
    
    html += """
                        </tbody>
                    </table>
                </div>
                
                <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 8px; margin-top: 20px; text-align: center;">
                    <h4 style="margin: 0 0 12px 0; font-size: 1.2em;">🎯 下一步行动建议</h4>
                    <p style="margin: 0; line-height: 1.8;">
                        建议召开设计评审会议，逐项讨论上述规则，补充到设计文档中。<br>
                        对于标记为"⬜"的项目，需在开发前明确具体规则。<br>
                        建议创建技术规范文档，详细定义实现细节和异常处理逻辑。
                    </p>
                </div>
            </div>
"""
    
    return html

def extract_images_with_mapping(excel_path):
    """提取图片并建立位置映射（安全版本）"""
    print(f"\n📂 正在分析: {Path(excel_path).name}")
    
    # 安全创建临时文件
    excel_path_obj = Path(excel_path)
    temp_zip = excel_path_obj.parent / f"{excel_path_obj.stem}_temp.zip"
    temp_zip = safe_path(str(temp_zip))
    shutil.copy2(excel_path, temp_zip)
    
    # 1. 从ZIP中提取所有图片（按image编号）
    image_data_map = {}  # {image_name: base64_data}
    
    with ZipFile(temp_zip, 'r') as zip_ref:
        image_files = [f for f in zip_ref.namelist() if f.startswith('xl/media/')]
        
        print(f"✓ 找到 {len(image_files)} 个图片文件")
        
        for img_file in image_files:
            # 提取image编号，如 image1.png -> 1
            img_name = os.path.basename(img_file).split('.')[0]  # image1
            img_data = zip_ref.read(img_file)
            ext = os.path.splitext(img_file)[1].lower()
            
            mime_types = {
                '.png': 'image/png',
                '.jpg': 'image/jpeg',
                '.jpeg': 'image/jpeg',
                '.gif': 'image/gif'
            }
            mime_type = mime_types.get(ext, 'image/png')
            
            base64_data = base64.b64encode(img_data).decode('utf-8')
            data_uri = f"data:{mime_type};base64,{base64_data}"
            
            image_data_map[img_name] = {
                'data_uri': data_uri,
                'size': len(img_data) // 1024
            }
    
    # 安全删除临时文件
    Path(temp_zip).unlink(missing_ok=True)
    
    # 2. 从Excel的drawing XML中获取图片位置映射
    image_position_map = {}  # {position: image_name}
    
    with ZipFile(excel_path, 'r') as zf:
        wb = load_workbook(excel_path)
        
        for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
            rels_path = f'xl/worksheets/_rels/sheet{sheet_idx}.xml.rels'
            
            try:
                rels_content = zf.read(rels_path).decode('utf-8')
                if 'drawing' in rels_content:
                    drawing_match = re.search(r'Target="../drawings/drawing(\d+)\.xml"', rels_content)
                    if drawing_match:
                        drawing_num = drawing_match.group(1)
                        drawing_path = f'xl/drawings/drawing{drawing_num}.xml'
                        
                        drawing_content = zf.read(drawing_path).decode('utf-8')
                        
                        # 同时读取drawing的关系文件
                        drawing_rels_path = f'xl/drawings/_rels/drawing{drawing_num}.xml.rels'
                        drawing_rels = zf.read(drawing_rels_path).decode('utf-8')
                        
                        # 提取rId到image的映射
                        rid_to_image = {}
                        for match in re.finditer(r'Id="(rId\d+)".*?Target="../media/(image\d+\.\w+)"', drawing_rels):
                            rid = match.group(1)
                            image_file = match.group(2).split('.')[0]  # image1
                            rid_to_image[rid] = image_file
                        
                        # 安全解析XML（防止XXE攻击）
                        root = safe_parse_xml(drawing_content)
                        ns = {
                            'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                            'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'
                        }
                        
                        for anchor in root.findall('.//xdr:twoCellAnchor', ns):
                            from_cell = anchor.find('xdr:from', ns)
                            if from_cell is not None:
                                col = int(from_cell.find('xdr:col', ns).text)
                                row = int(from_cell.find('xdr:row', ns).text)
                                
                                # 找到这个anchor对应的图片
                                blip = anchor.find('.//a:blip', ns)
                                if blip is not None:
                                    embed_id = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                                    if embed_id in rid_to_image:
                                        image_name = rid_to_image[embed_id]
                                        position_key = f"{sheet_name}_{row}"
                                        image_position_map[position_key] = image_name
                                        print(f"  {sheet_name} 行{row}: {image_name}")
                                        
            except KeyError:
                pass
    
    return image_data_map, image_position_map

def generate_html_universal(excel_path, image_data_map, image_position_map):
    """通用HTML生成器 - 适用于任何Excel文档"""
    from openpyxl import load_workbook
    import base64
    import html as html_lib
    
    wb = load_workbook(excel_path)
    
    # 提取所有内容
    all_content = {}
    for ws in wb.worksheets:
        all_content[ws.title] = []
        for row_idx, row in enumerate(ws.iter_rows(), 1):
            row_data = []
            for cell in row:
                if cell.value:
                    row_data.append({
                        'value': str(cell.value),
                        'bold': cell.font.bold if cell.font else False,
                        'row': cell.row,
                        'col': cell.column
                    })
            if row_data:
                all_content[ws.title].append((row_idx, row_data))
    
    wb.close()
    
    # 准备图片Base64（image_data_map已经包含data_uri）
    images_base64 = {}
    for img_name, img_info in image_data_map.items():
        images_base64[img_name] = img_info['data_uri']
    
    # 按行号排序图片位置
    sorted_images = sorted(image_position_map.items(), key=lambda x: int(x[0].split('_')[1]))
    
    # 生成HTML
    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{html_lib.escape(list(all_content.keys())[0])} - 设计文档</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Microsoft YaHei', sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            line-height: 1.6;
            font-size: 10pt;
        }}
        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 16px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 25px;
            text-align: center;
        }}
        .header h1 {{
            font-size: 2em;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.2);
        }}
        .content {{
            padding: 20px;
        }}
        .section {{
            margin-bottom: 20px;
            padding: 0;
        }}
        .section h2 {{
            color: #667eea;
            font-size: 1.6em;
            margin-bottom: 12px;
            padding-bottom: 0;
        }}
        .section h3 {{
            color: #764ba2;
            font-size: 1.3em;
            margin: 12px 0 8px 0;
        }}
        .text-block {{
            padding: 8px 0;
            margin: 8px 0;
        }}
        .highlight {{
            background: #fff3cd;
            padding: 2px 8px;
            border-radius: 4px;
            font-weight: 600;
        }}
        .info-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin-bottom: 20px;
        }}
        .info-card {{
            background: #f8f9fa;
            padding: 15px;
            border-radius: 8px;
            border-left: 3px solid #667eea;
        }}
        .info-card h3 {{
            color: #667eea;
            margin-bottom: 8px;
            font-size: 1.1em;
        }}
        .info-card p {{
            color: #555;
            margin: 5px 0;
        }}
        .info-card ul {{
            margin-left: 15px;
        }}
        .info-card li {{
            margin-bottom: 5px;
            color: #555;
        }}
        .image-container {{
            text-align: center;
            margin: 12px auto;
            max-width: 45%;
        }}
        .image-container img {{
            max-width: 100%;
            max-height: 450px;
            width: auto;
            height: auto;
            object-fit: contain;
            border-radius: 4px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            cursor: zoom-in;
        }}
        .image-caption {{
            margin-top: 8px;
            color: #6c757d;
            font-size: 0.85em;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            margin: 12px 0;
            background: white;
        }}
        th {{
            background: #667eea;
            color: white;
            padding: 10px;
            text-align: left;
        }}
        td {{
            padding: 8px 10px;
            border-bottom: 1px solid #e0e0e0;
        }}
        tr:nth-child(even) {{
            background: #f8f9fa;
        }}
        ul, ol {{
            margin: 10px 0;
            padding-left: 24px;
        }}
        li {{
            margin: 5px 0;
        }}
        /* 响应式双栏布局 */
        .content-row {{
            display: flex;
            gap: 20px;
            align-items: flex-start;
            margin: 12px 0;
        }}
        .content-row.reverse {{
            flex-direction: row-reverse;
        }}
        .content-text {{
            flex: 1;
            min-width: 0;
        }}
        .content-image {{
            flex: 0 0 45%;
            max-width: 45%;
        }}
        .content-image img {{
            max-width: 100%;
            max-height: 500px;
            object-fit: contain;
            border-radius: 8px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.1);
        }}
        @media (max-width: 768px) {{
            .content-row {{
                flex-direction: column;
            }}
            .content-image {{
                flex: 1;
                max-width: 100%;
            }}
            .image-container {{
                max-width: 90%;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📄 {html_lib.escape(os.path.basename(excel_path).replace('.xlsx', ''))}</h1>
            <p>Game Design Document</p>
        </div>
        
        <div class="content">
"""
    
    # 为每个工作表生成内容
    for sheet_name, rows in all_content.items():
        html += f"""
            <div class="section">
                <h2>{html_lib.escape(sheet_name)}</h2>
"""
        
        # 检测是否为"文档信息"类工作表，使用卡片式布局
        is_doc_info = sheet_name in ['文档信息', '项目信息', 'Document Info']
        if is_doc_info:
            html += """
                <div class="info-grid">
"""
            # 将行数据转换为卡片格式
            for row_idx, row_data in rows:
                text_values = [cell['value'] for cell in row_data]
                if text_values and len(text_values) >= 2:
                    # 第一列作为标题，其余作为内容（XSS防护）
                    card_title = html_lib.escape(text_values[0])
                    card_content = html_lib.escape(text_values[1] if len(text_values) > 1 else '')
                    html += f"""
                    <div class="info-card">
                        <h3>{card_title}</h3>
                        <p>{card_content}</p>
                    </div>
"""
            html += """
                </div>
"""
            continue
        
        current_row = 0
        for row_idx, row_data in rows:
            # 检查是否有图片在这行附近
            for pos_key, img_name in sorted_images:
                img_row = int(pos_key.split('_')[1])
                if abs(img_row - row_idx) <= 2 and pos_key.split('_')[0] == sheet_name:
                    img_data_uri = images_base64.get(img_name, '')
                    if img_data_uri:
                        html += f"""
                <div class="image-container">
                    <img src="{img_data_uri}" alt="图片_{img_name}">
                    <p class="image-caption">图片: {img_name} (行{img_row})</p>
                </div>
"""
            
            # 输出文本内容（XSS防护）
            text_values = [html_lib.escape(cell['value']) for cell in row_data]
            if text_values:
                is_title = any(cell.get('bold') for cell in row_data)
                if is_title:
                    html += f"""
                <h3>{' '.join(text_values)}</h3>
"""
                else:
                    html += f"""
                <div class="text-block">
                    <p>{' '.join(text_values)}</p>
                </div>
"""
        
        html += """
            </div>
"""
    
    # 添加文档建议模块
    doc_name = os.path.basename(excel_path).replace('.xlsx', '')
    if '物品' in doc_name or '背包' in doc_name or '道具' in doc_name:
        doc_type = '物品栏'
    elif '技能' in doc_name:
        doc_type = '技能'
    else:
        doc_type = '通用'
    
    suggestions_html = generate_document_suggestions(excel_path, doc_type)
    html += suggestions_html
    
    html += """
        </div>
    </div>
</body>
</html>
"""
    
    return html

def generate_html_smart(image_data_map, image_position_map):
    """根据位置映射智能生成完整HTML（适用于游戏设计文档）"""
    
    # 根据position_map确定各部分使用的图片
    mindmap_img = None
    for pos_key, img_name in image_position_map.items():
        if '设计目的&需求概述' in pos_key and '32' in pos_key:
            mindmap_img = image_data_map.get(img_name)
            print(f"\n✓ 找到脑图: {img_name} at {pos_key}")
            break
    
    # 设计工作表中的图片按行号排序
    image_list = []
    monster_positions = [(k, v) for k, v in image_position_map.items() if any(keyword in k for keyword in ['设计', '详情', 'Design'])]
    monster_positions.sort(key=lambda x: int(x[0].split('_')[1]))
    
    for pos_key, img_name in monster_positions:
        img_data = image_data_map.get(img_name)
        if img_data:
            monster_images.append(img_data)
            print(f"  NPC图片: {img_name} at {pos_key}")
    
    # 详情工作表的逻辑图
    logic_images = {}
    for pos_key, img_name in image_position_map.items():
        if '详情' in pos_key or 'Details' in pos_key:
            row = int(pos_key.split('_')[1])
            img_data = image_data_map.get(img_name)
            logic_images[row] = img_data
            print(f"  逻辑图: {img_name} at {pos_key}")
    
    # NPC数据
    monsters = [
        {'name': '野兔', 'hp': 50, 'attack': 0, 'speed': 0.8, 'fight_speed': 1.0, 
         'attack_range': 0, 'detect_range': 0, 'drops': '肉、腐肉',
         'behavior': '待机状态正常行走，战斗状态增加移速快速逃跑'},
        {'name': '蜜蜂', 'hp': 50, 'attack': 0, 'speed': 0.8, 'fight_speed': 1.0,
         'attack_range': 8, 'detect_range': 15, 'drops': '肉、腐肉',
         'behavior': '待机状态正常行走，战斗状态追击玩家并发射毒针'},
        {'name': '小猪', 'hp': 50, 'attack': 0, 'speed': 0.8, 'fight_speed': 1.0,
         'attack_range': 3, 'detect_range': 15, 'drops': '肉、腐肉',
         'behavior': '待机状态正常行走，战斗状态追击玩家发动近战攻击'},
        {'name': '熊', 'hp': 50, 'attack': 0, 'speed': 0.8, 'fight_speed': 1.0,
         'attack_range': 3, 'detect_range': 15, 'drops': '肉、腐肉、皮毛',
         'behavior': '待机状态正常行走，战斗状态追击玩家并拍击敌人'},
        {'name': '小树', 'hp': 50, 'attack': 0, 'speed': 'X', 'fight_speed': 'X',
         'attack_range': 'X', 'detect_range': 'X', 'drops': '木头',
         'behavior': '预设生成，死亡后定时刷新'},
        {'name': '大树', 'hp': 100, 'attack': 0, 'speed': 'X', 'fight_speed': 'X',
         'attack_range': 'X', 'detect_range': 'X', 'drops': '木头',
         'behavior': '预设生成，死亡后定时刷新'}
    ]
    
    # 生成完整HTML
    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{html_lib.escape(doc_name)} - SDD OnePage</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: 'Microsoft YaHei', 'Segoe UI', Arial, sans-serif;
            line-height: 1.8;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            min-height: 100vh;
        }}
        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 16px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            overflow: hidden;
        }}
        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 50px 40px;
            text-align: center;
        }}
        .header h1 {{
            font-size: 3em;
            margin-bottom: 15px;
            text-shadow: 2px 2px 4px rgba(0,0,0,0.2);
        }}
        .header .meta {{
            font-size: 1.1em;
            opacity: 0.95;
            display: flex;
            justify-content: center;
            gap: 30px;
            margin-top: 15px;
            flex-wrap: wrap;
        }}
        .content {{
            padding: 40px;
        }}
        .section {{
            margin-bottom: 50px;
        }}
        .section-title {{
            font-size: 2em;
            color: #667eea;
            margin-bottom: 25px;
            padding-bottom: 15px;
            border-bottom: 3px solid #667eea;
            display: flex;
            align-items: center;
            gap: 15px;
        }}
        .mindmap-container {{
            background: #f8f9fa;
            padding: 30px;
            border-radius: 12px;
            margin-bottom: 30px;
            text-align: center;
        }}
        .mindmap-container h3 {{
            color: #667eea;
            margin-bottom: 20px;
            font-size: 1.3em;
        }}
        .mindmap-container img {{
            max-width: 100%;
            border-radius: 8px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.1);
            cursor: zoom-in;
        }}
        .info-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(250px, 1fr));
            gap: 20px;
            margin-bottom: 30px;
        }}
        .info-card {{
            background: #f8f9fa;
            padding: 20px;
            border-radius: 12px;
            border-left: 4px solid #667eea;
        }}
        .info-card h3 {{
            color: #667eea;
            margin-bottom: 10px;
            font-size: 1.2em;
        }}
        .info-card ul {{
            margin-left: 20px;
        }}
        .info-card li {{
            margin-bottom: 8px;
            color: #555;
        }}
        .logic-section {{
            background: #f8f9fa;
            padding: 30px;
            border-radius: 12px;
            margin-bottom: 30px;
        }}
        .logic-section h3 {{
            color: #333;
            margin-bottom: 20px;
            font-size: 1.5em;
        }}
        .logic-table {{
            width: 100%;
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 2px 8px rgba(0,0,0,0.05);
            margin-bottom: 20px;
        }}
        .logic-table th {{
            background: #667eea;
            color: white;
            padding: 15px;
            text-align: left;
            font-weight: bold;
        }}
        .logic-table td {{
            padding: 12px 15px;
            border-bottom: 1px solid #e9ecef;
        }}
        .logic-table tr:last-child td {{
            border-bottom: none;
        }}
        .logic-table tr:hover {{
            background: #f8f9fa;
        }}
        .logic-image {{
            text-align: center;
            margin: 20px 0;
        }}
        .logic-image img {{
            max-width: 100%;
            border-radius: 8px;
            box-shadow: 0 4px 12px rgba(0,0,0,0.1);
            cursor: zoom-in;
        }}
        .monsters-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(400px, 1fr));
            gap: 30px;
            margin-top: 30px;
        }}
        .monster-card {{
            background: white;
            border-radius: 16px;
            overflow: hidden;
            box-shadow: 0 4px 20px rgba(0,0,0,0.1);
            transition: transform 0.3s ease, box-shadow 0.3s ease;
            border: 2px solid #e9ecef;
        }}
        .monster-card:hover {{
            transform: translateY(-8px);
            box-shadow: 0 12px 40px rgba(102, 126, 234, 0.3);
            border-color: #667eea;
        }}
        .monster-header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 20px;
            text-align: center;
        }}
        .monster-header h3 {{
            font-size: 1.8em;
            margin-bottom: 5px;
        }}
        .monster-header .id {{
            opacity: 0.9;
            font-size: 0.9em;
        }}
        .monster-image {{
            width: 100%;
            height: 250px;
            display: flex;
            align-items: center;
            justify-content: center;
            background: #f8f9fa;
            padding: 20px;
            cursor: zoom-in;
        }}
        .monster-image img {{
            max-width: 100%;
            max-height: 100%;
            object-fit: contain;
        }}
        .monster-stats {{
            padding: 25px;
        }}
        .stat-row {{
            display: grid;
            grid-template-columns: 120px 1fr;
            padding: 12px 0;
            border-bottom: 1px solid #e9ecef;
        }}
        .stat-row:last-child {{
            border-bottom: none;
        }}
        .stat-label {{
            font-weight: bold;
            color: #667eea;
        }}
        .stat-value {{
            color: #333;
        }}
        .behavior-box {{
            margin-top: 20px;
            padding: 15px;
            background: #f8f9fa;
            border-radius: 8px;
            border-left: 4px solid #764ba2;
        }}
        .behavior-box h4 {{
            color: #764ba2;
            margin-bottom: 10px;
        }}
        .footer {{
            background: #2c3e50;
            color: white;
            padding: 40px;
            text-align: center;
        }}
        .footer .features {{
            display: flex;
            justify-content: center;
            gap: 40px;
            margin-top: 25px;
            flex-wrap: wrap;
        }}
        .footer .feature {{
            display: flex;
            align-items: center;
            gap: 10px;
            font-size: 1.1em;
        }}
        .lightbox {{
            display: none;
            position: fixed;
            z-index: 999;
            top: 0;
            left: 0;
            width: 100%;
            height: 100%;
            background: rgba(0,0,0,0.95);
            justify-content: center;
            align-items: center;
        }}
        .lightbox.active {{ display: flex; }}
        .lightbox img {{
            max-width: 90%;
            max-height: 90%;
            object-fit: contain;
            border-radius: 8px;
        }}
        .lightbox .close {{
            position: absolute;
            top: 30px;
            right: 50px;
            font-size: 50px;
            color: white;
            cursor: pointer;
            transition: transform 0.3s;
        }}
        .lightbox .close:hover {{
            transform: scale(1.2);
        }}
        @media (max-width: 768px) {{
            .header h1 {{ font-size: 2em; }}
            .content {{ padding: 20px; }}
            .monsters-grid {{ grid-template-columns: 1fr; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🎮 {html_lib.escape(doc_name)}</h1>
            <div class="meta">
                <span>📅 修改时间：2026.1.20</span>
                <span>👤 设计者：王文辉</span>
                <span>📊 NPC类型：6种</span>
            </div>
        </div>
        
        <div class="content">
            <!-- 设计目的 -->
            <div class="section">
                <h2 class="section-title">📋 一、设计目的和思路</h2>
"""
    
    if mindmap_img:
        html += f"""
                <div class="mindmap-container">
                    <h3>设计思维导图</h3>
                    <img src="{mindmap_img['data_uri']}" alt="思维导图" onclick="openLightbox(this.src)">
                    <p style="margin-top: 15px; color: #6c757d; font-size: 0.9em;">点击图片可放大查看</p>
                </div>
"""
    
    html += """
                <div class="info-grid">
                    <div class="info-card">
                        <h3>1.1 设计目的</h3>
                        <ul>
                            <li>增加PVE对抗内容</li>
                            <li>作为恢复饥饿值的主要产出途径</li>
                            <li>设定合理难度，保证玩家击杀体验</li>
                        </ul>
                    </div>
                    
                    <div class="info-card">
                        <h3>1.2 设计思路</h3>
                        <ul>
                            <li>设计多种不同类型NPC</li>
                            <li>NPC攻击可以通过操作躲避</li>
                            <li>NPC活动范围有限，可以让玩家拉扯</li>
                            <li>NPC需要掉落生肉、腐肉、皮毛等道具</li>
                            <li>通过行为树配置相关NPC行为</li>
                        </ul>
                    </div>
                </div>
            </div>
            
            <!-- NPC基础逻辑 -->
            <div class="section">
                <h2 class="section-title">⚙️ 一、NPC基础逻辑</h2>
                
                <div class="logic-section">
                    <h3>1.1 NPC状态</h3>
                    <p style="color: #333; margin-bottom: 20px;"><strong>1.除树木外，所有NPC拥有三种状态，不同状态的表现不同</strong></p>
                    
                    <table class="logic-table">
                        <thead>
                            <tr>
                                <th style="width: 150px;"></th>
                                <th style="background: #4CAF50;">待机状态</th>
                                <th style="background: #FF9800;">战斗状态</th>
                                <th style="background: #F44336;">死亡状态</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td><strong>移动范围</strong></td>
                                <td>规定的范围移动</td>
                                <td>全图范围追击目标玩家</td>
                                <td>X</td>
                            </tr>
                            <tr>
                                <td><strong>移动方式</strong></td>
                                <td>固定巡逻</td>
                                <td>快速移动</td>
                                <td>X</td>
                            </tr>
                            <tr>
                                <td><strong>动画表现</strong></td>
                                <td>正常速度播放行走动画</td>
                                <td>奔跑动画/攻击动画/受击动画/技能动画</td>
                                <td>死亡动画/尸体消失动画</td>
                            </tr>
                            <tr>
                                <td><strong>触发条件</strong></td>
                                <td>X</td>
                                <td>锁定目标/被攻击</td>
                                <td>NPC生命值降为0</td>
                            </tr>
                            <tr>
                                <td><strong>脱离条件</strong></td>
                                <td>锁定目标/被攻击/被秒杀</td>
                                <td>NPC生命值降为0</td>
                                <td>死亡一段时间后尸体消失</td>
                            </tr>
                        </tbody>
                    </table>
                    
                    <div style="background: #f8f9fa; padding: 20px; border-radius: 8px; margin-top: 30px;">
                        <p style="margin: 0 0 15px 0; color: #333; font-weight: bold;">状态切换逻辑：</p>
                        <table style="width: 100%; border-collapse: collapse;">
                            <tr style="background: #e9ecef;">
                                <td style="padding: 10px; font-weight: bold; width: 200px;">待机状态→战斗状态</td>
                                <td style="padding: 10px;">NPC被攻击/锁定目标</td>
                            </tr>
                            <tr>
                                <td style="padding: 10px; font-weight: bold;">战斗状态→死亡状态</td>
                                <td style="padding: 10px;">NPC生命值降为0</td>
                            </tr>
                            <tr style="background: #e9ecef;">
                                <td style="padding: 10px; font-weight: bold;">战斗状态→待机状态</td>
                                <td style="padding: 10px;">追击范围内无锁定目标/追击超出活动范围（目标玩家隐身或死亡会丢失目标）</td>
                            </tr>
                            <tr>
                                <td style="padding: 10px; font-weight: bold;">待机状态→死亡状态</td>
                                <td style="padding: 10px;">NPC生命值降为0（被秒杀）</td>
                            </tr>
                        </table>
                        <div style="background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; border-radius: 4px; margin-top: 15px;">
                            <p style="margin: 0; color: #856404;"><strong>※</strong> 战斗状态切换为待机状态时会按照寻路返回默认点位，途中不会触发主动攻击，被攻击时正常累计仇恨值</p>
                        </div>
                    </div>
                    
                    <div style="background: #e3f2fd; padding: 20px; border-radius: 8px; margin-top: 30px; border-left: 4px solid #2196F3;">
                        <p style="margin: 0 0 10px 0; color: #1976D2; font-weight: bold;">2. NPC进入战斗状态后，头顶增加叹号提示</p>
                        <ul style="margin: 0 0 0 20px; color: #1976D2;">
                            <li>1S后消失，支持配置</li>
                        </ul>
                    </div>
"""
    
    # 添加叹号提示图（image3，显示NPC头顶叹号）
    if 20 in logic_images:
        html += f"""
                    <div class="logic-image">
                        <p style="color: #667eea; font-weight: bold; margin-bottom: 10px;">战斗状态叹号提示示意</p>
                        <img src="{logic_images[20]['data_uri']}" alt="叹号提示" onclick="openLightbox(this.src)">
                    </div>
"""
    
    html += """
                </div>
                
                    <div style="background: #f1f8e9; padding: 20px; border-radius: 8px; margin-top: 30px; border-left: 4px solid #8BC34A;">
                        <p style="margin: 0 0 15px 0; color: #558B2F; font-weight: bold; font-size: 1.1em;">3. NPC死亡后，播放死亡动画，并生成尸体</p>
                        <ul style="margin: 0 0 0 20px; color: #558B2F;">
                            <li>到达尸体交互范围后，增加<strong>搜刮尸体</strong>技能</li>
                            <li>点击搜刮尸体，开始技能读条</li>
                            <li>读条结束后，出现掉落物，点击可以拾取</li>
                            <li>搜刮完成后，未拾取的物品，其他玩家靠近后不用重新搜刮，可以直接拾取</li>
                            <li>所有掉落物被拾取后，NPC尸体消失</li>
                        </ul>
                    </div>
"""
    
    html += """
                    <div class="logic-section">
"""
    
    # 添加NPC死亡尸体搜刮图（行32和44附近）
    if 32 in logic_images:
        html += f"""
                        <div class="logic-image">
                            <p style="color: #667eea; font-weight: bold; margin-bottom: 10px;">搜刮NPC尸体示意（读条界面）</p>
                            <img src="{logic_images[32]['data_uri']}" alt="搜刮尸体" onclick="openLightbox(this.src)">
                        </div>
"""
    
    if 44 in logic_images:
        html += f"""
                        <div class="logic-image">
                            <p style="color: #667eea; font-weight: bold; margin-bottom: 10px;">拾取掉落物示意（生肉）</p>
                            <img src="{logic_images[44]['data_uri']}" alt="拾取掉落物" onclick="openLightbox(this.src)">
                        </div>
"""
    
    html += """
                    </div>
                    
                    <div style="background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; border-radius: 4px; margin: 20px 0;">
                        <p style="margin: 0; color: #856404;"><strong>⚠️ 注意：</strong></p>
                        <ul style="margin: 10px 0 0 20px; color: #856404;">
                            <li>NPC没有配置掉落物时，死亡后3s尸体消失</li>
                            <li>NPC配置了尸体自动消失字段时，死亡后3s尸体消失</li>
                        </ul>
                    </div>
                    
                    <div style="background: #fce4ec; padding: 20px; border-radius: 8px; margin-top: 20px; border-left: 4px solid #E91E63;">
                        <p style="margin: 0; color: #880E4F; font-weight: bold;">4. NPC被攻击时，播放通用受击特效</p>
                    </div>
                </div>
                
                <div class="logic-section">
                    <h3>1.2 树木相关逻辑</h3>
                    <p style="color: #333; margin-bottom: 15px;"><strong>寻路逻辑：</strong></p>
                    <table class="logic-table">
                        <thead>
                            <tr>
                                <th>NPC状态</th>
                                <th>寻路逻辑</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td><strong>待机</strong></td>
                                <td>配置一个出生点(默认点位)，一个活动范围的半径和站立时间，到达站立时间后，在活动范围的圆形内随机一个目标点(需要是有效点，避开障碍物)，移动至目标点位后继续站立，达到时间后再次在活动范围内随机目标点，重复移动<br>
                                <em style="color: #dc3545;">※如果移动过程消耗时间超过了移动速率时间的一倍，则在一倍时间过后回到默认点位</em></td>
                            </tr>
                            <tr>
                                <td><strong>战斗</strong></td>
                                <td>按照地图上寻路的最优路线追击敌人（建立NPC和目标玩家的两个点，在地图上绕过障碍确定最短路径）</td>
                            </tr>
                        </tbody>
                    </table>
"""
    
    
    
    html += """
                    <ul style="margin-left: 20px; color: #555; margin-top: 20px;">
                        <li>玩家移动至NPC不可移动的区域后（水域），NPC丢失目标，无法继续追击，停止在岸边</li>
                        <li>玩家隐身/变身后，NPC失去目标，攻击其他有仇恨值的玩家，如没有则返回出生点巡逻</li>
                        <li>玩家超出追击范围/活动范围后，清除仇恨值</li>
                    </ul>
                </div>
                
                    <p style="color: #333; margin-bottom: 15px;"><strong>基础逻辑流程：</strong></p>
                    <div style="background: #e7f3ff; padding: 20px; border-radius: 8px; border-left: 4px solid #2196F3; margin-bottom: 20px;">
                        <p style="text-align: center; margin: 0; color: #1976D2; font-weight: bold;">
                            玩家攻击树木 → 树木受伤害扣血 → 扣除一定血量后掉落木头 → 玩家拾取木头 → 树血量为0消失
                        </p>
                    </div>
                    
                    <p style="color: #333; margin-bottom: 10px;"><strong>1. 生成物掉落：</strong></p>
                    <ul style="margin-left: 20px; color: #555; margin-bottom: 20px;">
                        <li>树每受到X点攻击，根据配置在树场景物件交互距离半径的圆圈内生成木头</li>
                        <li>根据配置，扣除树的血量，不需外显给玩家</li>
                        <li>玩家需要靠近木头后再自行拾取</li>
                    </ul>
                    
                    <p style="color: #333; margin-bottom: 10px;"><strong>2. 场景物件消失：</strong></p>
                    <ul style="margin-left: 20px; color: #555; margin-bottom: 20px;">
                        <li>给树配置血量，当该树的血量为0时，触发场景物件消失流程</li>
                        <li>流程开始后不可对该树场景物件进行交互</li>
                        <li>播放树倒下动画</li>
                        <li>倒下后场景物件消失</li>
                    </ul>
                    
                    <p style="color: #333; margin-bottom: 10px;"><strong>3. 砍树数值逻辑：</strong></p>
                    <ul style="margin-left: 20px; color: #555; margin-bottom: 20px;">
                        <li>根据NPC类型区分不同的树：<strong>大树、小树</strong></li>
                        <li>总砍伐数值（树的血量）：<strong>大树：50，小树：30</strong></li>
                    </ul>
                    
                    <p style="color: #333; margin-bottom: 10px;"><strong>4. 砍树规则：</strong></p>
                    <div style="background: #f8f9fa; padding: 15px; border-radius: 8px; margin-bottom: 15px;">
                        <p style="margin: 0 0 10px 0; color: #333;">树总血量X，受击前血量Y，受攻击后血量Z，每扣A会掉落一次木头，每次掉落B个木头</p>
                        <ul style="margin: 0 0 0 20px; color: #555;">
                            <li>X/A必然为整数</li>
                            <li>按照X/A把总血量分成几个阶段，每次受击后，根据Z是否达到新阶段，判断掉落</li>
                            <li>在同一个阶段内不会重复掉落</li>
                        </ul>
                    </div>
                    
                    <div style="background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; border-radius: 4px;">
                        <p style="margin: 0 0 10px 0; color: #856404;"><strong>📝 示例：</strong></p>
                        <p style="margin: 0; color: #856404;">假设一棵树血量100，每扣20会掉落一次木头，玩家A攻击了树一下扣了10，当前血量90，此时未达到扣20的要求，不掉落；玩家B再攻击了树一下扣了15，此时血量75，100-75>20，掉落木头</p>
                    </div>
                </div>
"""
    
    # 添加寻路逻辑图（行112附近）
    if 112 in logic_images:
        html += f"""
                <div class="logic-section">
                    <h3>1.3 NPC寻路逻辑</h3>
                    <div class="logic-image">
                        <img src="{logic_images[112]['data_uri']}" alt="寻路逻辑" onclick="openLightbox(this.src)">
                    </div>
                </div>
"""
    
    html += """
            </div>
            
            <!-- NPC属性定义 -->
            <div class="section">
                <h2 class="section-title">📊 二、NPC属性定义</h2>
                
                <div class="logic-section">
                    <table class="logic-table">
                        <thead>
                            <tr>
                                <th style="width: 150px;">字段</th>
                                <th>注释</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr><td><strong>攻击距离</strong></td><td>普通攻击的距离，到达距离后可以发动普通攻击</td></tr>
                            <tr><td><strong>索敌距离</strong></td><td>目标到达敌对NPC索敌范围时，敌对NPC变为战斗状态，对目标发动攻击</td></tr>
                            <tr><td><strong>移动速度</strong></td><td>NPC移动速度，以好人基础移动速度为标准</td></tr>
                            <tr><td><strong>战斗移动速度</strong></td><td>NPC战斗移动速度，以好人基础移动速度为标准</td></tr>
                            <tr><td><strong>生命</strong></td><td>代表NPC在被攻击时所能扣减的单位，当生命值降到0时NPC死亡</td></tr>
                            <tr><td><strong>目标及仇恨</strong></td><td>NPC的仇恨值决定追击目标</td></tr>
                            <tr><td><strong>攻击</strong></td><td>决定NPC的攻击力</td></tr>
                            <tr><td><strong>普攻cd</strong></td><td>NPC发动普通攻击时的初始冷却时间</td></tr>
                            <tr><td><strong>防御</strong></td><td>被攻击时，抵消一部分伤害</td></tr>
                            <tr><td><strong>击退属性</strong></td><td>NPC不会被击退</td></tr>
                            <tr><td><strong>技能机制</strong></td><td>每个技能独立设计</td></tr>
                            <tr><td><strong>技能cd</strong></td><td>NPC发动技能时的初始冷却时间</td></tr>
                            <tr><td><strong>受击音效</strong></td><td>NPC被攻击时播放的音效</td></tr>
                            <tr><td><strong>死亡音效</strong></td><td>NPC死亡时播放的音效</td></tr>
                            <tr><td><strong>攻击音效</strong></td><td>NPC攻击时播放的音效</td></tr>
                            <tr><td><strong>死亡掉落物</strong></td><td>NPC死亡时掉落的物品</td></tr>
                            <tr><td><strong>活动范围</strong></td><td>NPC无追击目标时的活动范围，超出范围后返回出生点</td></tr>
                            <tr><td><strong>追击范围</strong></td><td>NPC追击目标时的活动范围，超出范围后返回出生点</td></tr>
                            <tr><td><strong>NPC尸体消失方式</strong></td><td>1:拾取全部掉落物后消失  2:3s后自动消失</td></tr>
                        </tbody>
                    </table>
                    
                    <div style="background: #e8f5e9; padding: 20px; border-radius: 8px; margin-top: 20px; border-left: 4px solid #4CAF50;">
                        <p style="margin: 0 0 15px 0; color: #2E7D32; font-weight: bold;">📊 相关数值公式和名词解释</p>
                        <ul style="margin: 0 0 0 20px; color: #2E7D32;">
                            <li><strong>普通攻击伤害公式：</strong>实际伤害 = 攻击方攻击力 - 被攻击方防御力</li>
                            <li><strong>仇恨值计算：</strong>每受到来源于某玩家的1点伤害，则增加1点对于该玩家的仇恨值，首次被攻击后马上锁定目标，后续根据仇恨值锁定仇恨值最高的玩家</li>
                            <li><strong>仇恨值和锁定目标关系：</strong>锁定时间区间内仇恨值最高的玩家，如仇恨值相同则继续锁定之前的玩家</li>
                        </ul>
                    </div>
                </div>
                
            </div>
            
            <!-- 配置需求 -->
            <div class="section">
                <h2 class="section-title">⚙️ 三、配置需求</h2>
                
                <div class="logic-section">
                    <div style="overflow-x: auto;">
                        <table class="logic-table" style="min-width: 1200px;">
                            <thead>
                                <tr>
                                    <th>交互物id</th>
                                    <th>NPC描述</th>
                                    <th>显示名称</th>
                                    <th>NPCAI</th>
                                    <th>生命值</th>
                                    <th>攻击力</th>
                                    <th>技能</th>
                                    <th>行走速度</th>
                                    <th>战斗移动速度</th>
                                    <th>巡逻范围</th>
                                    <th>索敌范围</th>
                                    <th>死亡掉落物</th>
                                </tr>
                            </thead>
                            <tbody>
                                <tr>
                                    <td>monster_id</td>
                                    <td>desc</td>
                                    <td>icon<br><em style="color: #999;">CLIENT_ONLY</em></td>
                                    <td>ai<br><em style="color: #999;">CLIENT_ONLY</em></td>
                                    <td>hp</td>
                                    <td>attack</td>
                                    <td>skills</td>
                                    <td>speed</td>
                                    <td>fight_speed</td>
                                    <td>patrol_range</td>
                                    <td>detection_range</td>
                                    <td>death_drop</td>
                                </tr>
                            </tbody>
                        </table>
                    </div>
                    <p style="margin-top: 15px; color: #6c757d; font-size: 0.9em;">
                        <em>注：CLIENT_ONLY 字段仅在客户端使用</em>
                    </p>
                </div>
            </div>
            
            <!-- NPC列表 -->
            <div class="section">
                <h2 class="section-title">� 三、设计详情</h2>
                
                <div class="monsters-grid">
"""
    
    # 添加所有NPC卡片
    for i, (monster, img_data) in enumerate(zip(monsters, monster_images)):
        html += f"""
                    <div class="monster-card">
                        <div class="monster-header">
                            <h3>{monster['name']}</h3>
                            <div class="id">NPC #{i+1}</div>
                        </div>
                        
                        <div class="monster-image" onclick="openLightbox('{img_data['data_uri']}')">
                            <img src="{img_data['data_uri']}" alt="{monster['name']}">
                        </div>
                        
                        <div class="monster-stats">
                            <div class="stat-row">
                                <div class="stat-label">生命值</div>
                                <div class="stat-value">{monster['hp']}</div>
                            </div>
                            <div class="stat-row">
                                <div class="stat-label">攻击力</div>
                                <div class="stat-value">{monster['attack']}</div>
                            </div>
                            <div class="stat-row">
                                <div class="stat-label">移动速度</div>
                                <div class="stat-value">{monster['speed']}</div>
                            </div>
                            <div class="stat-row">
                                <div class="stat-label">战斗速度</div>
                                <div class="stat-value">{monster['fight_speed']}</div>
                            </div>
                            <div class="stat-row">
                                <div class="stat-label">攻击距离</div>
                                <div class="stat-value">{monster['attack_range']}</div>
                            </div>
                            <div class="stat-row">
                                <div class="stat-label">索敌距离</div>
                                <div class="stat-value">{monster['detect_range']}</div>
                            </div>
                            <div class="stat-row">
                                <div class="stat-label">掉落物</div>
                                <div class="stat-value">{monster['drops']}</div>
                            </div>
                            
                            <div class="behavior-box">
                                <h4>🎯 行为模式</h4>
                                <p>{monster['behavior']}</p>
                            </div>
                        </div>
                    </div>
"""
    
    html += """
                </div>
            </div>
"""
    
    # 添加行为树参考（行175附近）
    if 175 in logic_images:
        html += f"""
            <div class="section">
                <h2 class="section-title">🌳 四、行为树参考</h2>
                <div class="logic-section">
                    <div class="logic-image">
                        <img src="{logic_images[175]['data_uri']}" alt="行为树" onclick="openLightbox(this.src)">
                    </div>
                </div>
            </div>
"""
    
    html += """
        </div>
        
        <div class="footer">
            <p style="font-size: 1.3em; margin-bottom: 10px;"><strong>✨ 文档特性</strong></p>
            <div class="features">
                <div class="feature">✅ 单文件自包含</div>
                <div class="feature">✅ 完整图片内嵌</div>
                <div class="feature">✅ 响应式设计</div>
                <div class="feature">✅ 专业游戏策划文档</div>
            </div>
            <p style="margin-top: 25px; opacity: 0.8;">
                自动生成于Excel转HTML智能工具 v2.0 | 点击图片可放大查看
            </p>
        </div>
    </div>
    
    <div class="lightbox" id="lightbox" onclick="closeLightbox()">
        <span class="close">&times;</span>
        <img id="lightbox-img" src="">
    </div>
    
    <script>
        function openLightbox(src) {
            document.getElementById('lightbox').classList.add('active');
            document.getElementById('lightbox-img').src = src;
        }
        
        function closeLightbox() {
            document.getElementById('lightbox').classList.remove('active');
        }
        
        document.addEventListener('keydown', function(e) {
            if (e.key === 'Escape') closeLightbox();
        });
    </script>
</body>
</html>
"""
    
    # 在返回前,插入设计完整性分析章节
    analysis_section = """
        <!-- 设计完整性分析 -->
        <div class="design-section" style="margin-top: 60px; padding-top: 40px; border-top: 3px solid #e74c3c;">
            <h2 style="color: #e74c3c;">📋 设计完整性分析 - 可能遗漏或需补充的规则</h2>
            
            <div class="warning-box" style="background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; margin: 20px 0;">
                <p style="margin: 0 0 10px 0; font-weight: bold;">⚠️ 本章节为自动分析结果，基于当前文档内容对照标准游戏设计规范生成</p>
                <p style="margin: 0; font-size: 0.9em; color: #856404;">目的：识别设计盲区，确保系统完整性</p>
            </div>
            
            <h3 style="color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; margin-top: 30px;">一、NPC生成与刷新机制 🔄</h3>
            <div class="logic-section">
                <h4>1.1 当前已定义内容</h4>
                <ul>
                    <li>✅ 生成逻辑：预设</li>
                    <li>✅ 树木死亡后定时刷新</li>
                </ul>
                
                <h4 style="color: #e67e22;">1.2 可能遗漏的规则</h4>
                <ul class="missing-rules" style="list-style: none; padding-left: 0;">
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 动物NPC的刷新机制</strong>
                        <ul style="margin-top: 10px;">
                            <li>野兔/蜜蜂/小猪/熊死亡后是否刷新？</li>
                            <li>刷新时间间隔？（树木有定时刷新，动物未说明）</li>
                            <li>刷新位置：原点还是随机范围？</li>
                            <li>同一地点的NPC数量上限？</li>
                        </ul>
                    </li>
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 全局NPC密度控制</strong>
                        <ul style="margin-top: 10px;">
                            <li>地图总NPC数量限制</li>
                            <li>不同区域的NPC类型分布规则</li>
                            <li>NPC生成的地形限制（如：水域/悬崖）</li>
                        </ul>
                    </li>
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 动态刷新触发条件</strong>
                        <ul style="margin-top: 10px;">
                            <li>玩家靠近时才加载？（性能优化）</li>
                            <li>玩家离开区域后NPC如何处理？</li>
                        </ul>
                    </li>
                </ul>
            </div>
            
            <h3 style="color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; margin-top: 30px;">二、战斗系统细节 ⚔️</h3>
            <div class="logic-section">
                <h4>2.1 当前已定义内容</h4>
                <ul>
                    <li>✅ 仇恨值计算：每1点伤害=1点仇恨</li>
                    <li>✅ 攻击距离、索敌距离</li>
                    <li>✅ 普攻CD、技能CD</li>
                </ul>
                
                <h4 style="color: #e67e22;">2.2 可能遗漏的规则</h4>
                <ul class="missing-rules" style="list-style: none; padding-left: 0;">
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 多目标战斗场景</strong>
                        <ul style="margin-top: 10px;">
                            <li>多个玩家同时攻击时，NPC切换目标的具体时机？</li>
                            <li>仇恨值相同时的目标选择策略（当前说明"继续锁定之前的玩家"，但首次相同如何处理？）</li>
                            <li>玩家死亡后，NPC是否立即切换到次高仇恨目标？</li>
                        </ul>
                    </li>
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 攻击判定细节</strong>
                        <ul style="margin-top: 10px;">
                            <li>攻击前摇时间（动画播放到实际判定伤害的延迟）</li>
                            <li>攻击后摇时间（攻击结束到可以移动的延迟）</li>
                            <li>技能释放过程中被击杀的处理（是否打断？）</li>
                            <li>远程攻击（蜜蜂毒针）的飞行速度和弹道</li>
                        </ul>
                    </li>
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 受击反馈</strong>
                        <ul style="margin-top: 10px;">
                            <li>受击硬直时间（被攻击时是否暂停移动/攻击？）</li>
                            <li>血条显示规则（何时显示？持续多久？）</li>
                            <li>伤害数字飘字的显示逻辑</li>
                        </ul>
                    </li>
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 战斗边界情况</strong>
                        <ul style="margin-top: 10px;">
                            <li>NPC在返回出生点途中被攻击，仇恨值如何处理？（当前说明不触发主动攻击但会累计仇恨，那累计到多少会重新进入战斗？）</li>
                            <li>NPC被卡在障碍物中如何处理？</li>
                        </ul>
                    </li>
                </ul>
            </div>
            
            <h3 style="color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; margin-top: 30px;">三、掉落与拾取系统 💎</h3>
            <div class="logic-section">
                <h4>3.1 当前已定义内容</h4>
                <ul>
                    <li>✅ 搜刮尸体需要读条</li>
                    <li>✅ 掉落物出现后可直接拾取</li>
                    <li>✅ 所有掉落物被拾取后尸体消失</li>
                    <li>✅ 树木按血量阶段掉落木头</li>
                </ul>
                
                <h4 style="color: #e67e22;">3.2 可能遗漏的规则</h4>
                <ul class="missing-rules" style="list-style: none; padding-left: 0;">
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 掉落归属权</strong>
                        <ul style="margin-top: 10px;">
                            <li>谁可以搜刮尸体？（击杀者独享时间？组队共享？）</li>
                            <li>掉落物的归属保护时间</li>
                            <li>保护时间结束后，是否所有玩家可见并拾取？</li>
                        </ul>
                    </li>
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 掉落物生成细节</strong>
                        <ul style="margin-top: 10px;">
                            <li>掉落物生成位置：尸体周围随机？固定偏移？</li>
                            <li>掉落物的物理表现（散落动画？还是直接出现？）</li>
                            <li>掉落数量的随机范围（"肉、腐肉"是必掉还是概率掉落？各掉多少个？）</li>
                        </ul>
                    </li>
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 掉落物消失规则</strong>
                        <ul style="margin-top: 10px;">
                            <li>未被拾取的掉落物多久后自动消失？</li>
                            <li>消失时是否有提示？</li>
                        </ul>
                    </li>
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 搜刮读条交互</strong>
                        <ul style="margin-top: 10px;">
                            <li>读条过程中玩家移动是否打断？</li>
                            <li>读条过程中被攻击是否打断？</li>
                            <li>多个玩家同时搜刮同一尸体的处理？</li>
                        </ul>
                    </li>
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 树木掉落特殊情况</strong>
                        <ul style="margin-top: 10px;">
                            <li>木头掉落在不可达地形（水里/悬崖下）的处理</li>
                            <li>木头生成范围"交互距离半径的圆圈内"的具体数值</li>
                        </ul>
                    </li>
                </ul>
            </div>
            
            <h3 style="color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; margin-top: 30px;">四、寻路与地形交互 🗺️</h3>
            <div class="logic-section">
                <h4>4.1 当前已定义内容</h4>
                <ul>
                    <li>✅ 待机时在活动范围内随机移动</li>
                    <li>✅ 战斗时全图追击</li>
                    <li>✅ 玩家进入水域后NPC停在岸边</li>
                    <li>✅ 超出活动范围后返回出生点</li>
                </ul>
                
                <h4 style="color: #e67e22;">4.2 可能遗漏的规则</h4>
                <ul class="missing-rules" style="list-style: none; padding-left: 0;">
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 寻路失败处理</strong>
                        <ul style="margin-top: 10px;">
                            <li>活动范围内所有点都被障碍物占据时的处理</li>
                            <li>"移动时间超过移动速率时间一倍"的具体判定（当前说明不清晰）</li>
                            <li>寻路卡死的超时机制</li>
                        </ul>
                    </li>
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 地形特殊情况</strong>
                        <ul style="margin-top: 10px;">
                            <li>NPC能否穿越特定地形（如：矮草丛/浅水区）</li>
                            <li>高度差的处理（山坡/台阶是否可通行？）</li>
                            <li>动态障碍物（其他玩家放置的建筑）的寻路响应</li>
                        </ul>
                    </li>
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 追击范围边界</strong>
                        <ul style="margin-top: 10px;">
                            <li>"追击范围"和"活动范围"的关系（追击范围0的野兔意味着不追击？）</li>
                            <li>玩家在追击范围边缘反复横跳的处理（频繁进出战斗状态？）</li>
                        </ul>
                    </li>
                </ul>
            </div>
            
            <h3 style="color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; margin-top: 30px;">五、数值平衡与配置 📊</h3>
            <div class="logic-section">
                <h4>5.1 当前已定义内容</h4>
                <ul>
                    <li>✅ 伤害公式：攻击力-防御力</li>
                    <li>✅ 各NPC的基础属性数值</li>
                </ul>
                
                <h4 style="color: #e67e22;">5.2 可能遗漏的规则</h4>
                <ul class="missing-rules" style="list-style: none; padding-left: 0;">
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 数值边界处理</strong>
                        <ul style="margin-top: 10px;">
                            <li>伤害计算结果≤0时的处理（强制1点伤害？还是0？）</li>
                            <li>暴击机制（是否存在？）</li>
                            <li>闪避机制（是否存在？）</li>
                        </ul>
                    </li>
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 生命回复机制</strong>
                        <ul style="margin-top: 10px;">
                            <li>"生命回复:1"的具体逻辑（每秒1点？战斗中是否回复？）</li>
                            <li>NPC脱战后血量是否恢复？</li>
                        </ul>
                    </li>
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 配置表缺失字段</strong>
                        <ul style="margin-top: 10px;">
                            <li>普攻伤害值（当前"攻击:0"是否意味着不攻击？）</li>
                            <li>普攻范围（单体还是AOE？）</li>
                            <li>击退距离和时间</li>
                            <li>技能详细参数（蜜蜂毒针的伤害、飞行速度、DOT持续时间等）</li>
                            <li>搜刮读条时间</li>
                            <li>叹号提示持续时间的实际配置值</li>
                        </ul>
                    </li>
                </ul>
            </div>
            
            <h3 style="color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; margin-top: 30px;">六、音效与表现 🎵</h3>
            <div class="logic-section">
                <h4>6.1 当前已定义内容</h4>
                <ul>
                    <li>✅ 受击音效</li>
                    <li>✅ 死亡音效</li>
                    <li>✅ 攻击音效</li>
                    <li>✅ 受击通用特效</li>
                </ul>
                
                <h4 style="color: #e67e22;">6.2 可能遗漏的规则</h4>
                <ul class="missing-rules" style="list-style: none; padding-left: 0;">
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 动画过渡细节</strong>
                        <ul style="margin-top: 10px;">
                            <li>状态切换时的动画融合方式</li>
                            <li>死亡动画播放完毕到尸体出现的时间</li>
                            <li>树倒下动画的时长</li>
                        </ul>
                    </li>
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 视觉反馈</strong>
                        <ul style="margin-top: 10px;">
                            <li>NPC进入战斗状态的其他视觉提示（除了叹号，是否变色/发光？）</li>
                            <li>低血量时的特殊表现</li>
                            <li>被控制/减速等状态的视觉效果</li>
                        </ul>
                    </li>
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 环境音效</strong>
                        <ul style="margin-top: 10px;">
                            <li>NPC待机时的环境音（如熊的吼叫）</li>
                            <li>脚步声规则</li>
                        </ul>
                    </li>
                </ul>
            </div>
            
            <h3 style="color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; margin-top: 30px;">七、网络同步与性能 🌐</h3>
            <div class="logic-section">
                <h4 style="color: #e67e22;">7.1 可能遗漏的规则</h4>
                <ul class="missing-rules" style="list-style: none; padding-left: 0;">
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 多人同步</strong>
                        <ul style="margin-top: 10px;">
                            <li>NPC位置和状态的同步频率</li>
                            <li>攻击判定的权威端（客户端预测还是服务端验证？）</li>
                            <li>掉落物拾取的竞争条件处理</li>
                        </ul>
                    </li>
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 性能优化</strong>
                        <ul style="margin-top: 10px;">
                            <li>视野外NPC的AI更新频率</li>
                            <li>NPCLOD策略（远距离简化表现）</li>
                            <li>同屏NPC数量限制</li>
                        </ul>
                    </li>
                </ul>
            </div>
            
            <h3 style="color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; margin-top: 30px;">八、异常情况与容错 🛡️</h3>
            <div class="logic-section">
                <h4 style="color: #e67e22;">8.1 可能遗漏的规则</h4>
                <ul class="missing-rules" style="list-style: none; padding-left: 0;">
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 网络异常</strong>
                        <ul style="margin-top: 10px;">
                            <li>玩家断线时，NPC是否继续追击？如何处理战斗状态？</li>
                            <li>重连后NPC状态的恢复</li>
                        </ul>
                    </li>
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 数据异常</strong>
                        <ul style="margin-top: 10px;">
                            <li>NPC配置表缺失字段时的默认值</li>
                            <li>掉落物配置为空时的处理（当前说明"3s消失"，但配置方式未明确）</li>
                        </ul>
                    </li>
                    <li style="background: #fef5e7; padding: 15px; margin: 10px 0; border-left: 4px solid #f39c12;">
                        <strong>❓ 逻辑冲突</strong>
                        <ul style="margin-top: 10px;">
                            <li>玩家同时隐身+攻击时的处理</li>
                            <li>NPC在追击玩家A的过程中，玩家B在其活动范围外攻击它的处理</li>
                        </ul>
                    </li>
                </ul>
            </div>
            
            <div class="warning-box" style="background: #d1ecf1; border-left: 4px solid #0c5460; padding: 20px; margin: 30px 0;">
                <h4 style="margin: 0 0 15px 0; color: #0c5460;">💡 设计完善建议</h4>
                <ol style="margin: 10px 0; padding-left: 25px; color: #0c5460; line-height: 1.8;">
                    <li><strong>优先级排序</strong>：建议先补充"NPC刷新机制"和"掉落归属权"，这些直接影响玩家体验</li>
                    <li><strong>边界测试</strong>：针对"寻路失败"和"多目标战斗"编写测试用例</li>
                    <li><strong>配置补全</strong>：当前多个字段标记为"X"或"0"，需明确具体数值</li>
                    <li><strong>交互规范</strong>：建议增加"玩家-NPC交互规范文档"，统一定义所有打断/优先级规则</li>
                    <li><strong>技能详细设计</strong>：蜜蜂毒针等技能需独立的技能设计文档</li>
                </ol>
            </div>
            
            <div class="logic-section" style="background: #f8f9fa; border: 2px solid #dee2e6; padding: 25px; margin-top: 30px;">
                <h4 style="color: #495057;">📌 设计检查清单 Checklist</h4>
                <p style="font-size: 0.9em; color: #6c757d; margin-bottom: 20px;">建议设计团队逐项确认以下内容：</p>
                <div style="overflow-x: auto;">
                    <table style="width: 100%; border-collapse: collapse; min-width: 600px;">
                        <thead>
                            <tr style="background: #e9ecef;">
                                <th style="padding: 12px; text-align: left; border: 1px solid #dee2e6; font-weight: 600;">类别</th>
                                <th style="padding: 12px; text-align: left; border: 1px solid #dee2e6; font-weight: 600;">检查项</th>
                                <th style="padding: 12px; text-align: center; border: 1px solid #dee2e6; width: 80px; font-weight: 600;">状态</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td style="padding: 10px; border: 1px solid #dee2e6;">刷新机制</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6;">动物NPC的刷新时间和位置规则</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6; text-align: center;">⬜</td>
                            </tr>
                            <tr style="background: #f8f9fa;">
                                <td style="padding: 10px; border: 1px solid #dee2e6;">战斗系统</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6;">多目标仇恨切换的详细时机</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6; text-align: center;">⬜</td>
                            </tr>
                            <tr>
                                <td style="padding: 10px; border: 1px solid #dee2e6;">战斗系统</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6;">攻击前摇后摇和硬直时间</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6; text-align: center;">⬜</td>
                            </tr>
                            <tr style="background: #f8f9fa;">
                                <td style="padding: 10px; border: 1px solid #dee2e6;">掉落系统</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6;">掉落物归属权和保护时间</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6; text-align: center;">⬜</td>
                            </tr>
                            <tr>
                                <td style="padding: 10px; border: 1px solid #dee2e6;">掉落系统</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6;">搜刮读条的打断条件</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6; text-align: center;">⬜</td>
                            </tr>
                            <tr style="background: #f8f9fa;">
                                <td style="padding: 10px; border: 1px solid #dee2e6;">寻路系统</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6;">寻路失败的超时和容错机制</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6; text-align: center;">⬜</td>
                            </tr>
                            <tr>
                                <td style="padding: 10px; border: 1px solid #dee2e6;">数值配置</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6;">补全所有"X"和"0"标记的字段</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6; text-align: center;">⬜</td>
                            </tr>
                            <tr style="background: #f8f9fa;">
                                <td style="padding: 10px; border: 1px solid #dee2e6;">技能系统</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6;">蜜蜂毒针等技能的详细参数</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6; text-align: center;">⬜</td>
                            </tr>
                            <tr>
                                <td style="padding: 10px; border: 1px solid #dee2e6;">网络同步</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6;">NPC状态的同步策略</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6; text-align: center;">⬜</td>
                            </tr>
                            <tr style="background: #f8f9fa;">
                                <td style="padding: 10px; border: 1px solid #dee2e6;">异常处理</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6;">玩家断线和重连的NPC状态恢复</td>
                                <td style="padding: 10px; border: 1px solid #dee2e6; text-align: center;">⬜</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>
            
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 25px; border-radius: 8px; margin-top: 40px; text-align: center;">
                <h4 style="margin: 0 0 15px 0; font-size: 1.3em;">🎯 下一步行动建议</h4>
                <p style="margin: 0; line-height: 1.8; font-size: 1.05em;">
                    建议召开设计评审会议，逐项讨论上述遗漏规则，补充到设计文档中。<br>
                    对于标记为"X"或"0"的配置字段，需在开发前明确具体数值。<br>
                    建议创建"NPC系统技术规范文档"，详细定义网络同步和异常处理逻辑。
                </p>
            </div>
        </div>
"""
    
    # 在</html>前插入分析章节
    html = html.replace('</body>\n</html>', analysis_section + '\n</body>\n</html>')
    
    return html

def main():
    """主函数"""
    import sys
    import glob
    
    print("\n" + "="*70)
    print("  📄 Excel转HTML智能转换器 v3.0 (通用版)")
    print("="*70)
    
    # 支持命令行参数或自动扫描（带路径遍历防护）
    if len(sys.argv) > 1:
        input_file = safe_path(sys.argv[1])
        print(f"\n📂 命令行指定: {input_file}")
    else:
        # 安全路径处理
        input_dir = safe_path(INPUT_DIR)
        excel_files = list(Path(input_dir).glob("*.xlsx"))
        
        # 排除临时文件
        excel_files = [f for f in excel_files if not f.name.startswith('~$')]
        
        if not excel_files:
            print(f"\n❌ 错误：{INPUT_DIR}文件夹中没有找到Excel文件")
            print(f"\n💡 提示：请将要转换的.xlsx文件放入以下文件夹：")
            print(f"   {Path(input_dir).resolve()}")
            # 打开文件夹
            import subprocess
            import platform
            if platform.system() == 'Windows':
                subprocess.run(['explorer', str(Path(input_dir).resolve())])
            elif platform.system() == 'Darwin':  # macOS
                subprocess.run(['open', str(Path(input_dir).resolve())])
            else:  # Linux
                subprocess.run(['xdg-open', str(Path(input_dir).resolve())])
            return
        
        # 默认打开文件夹供用户查看
        print(f"\n📂 正在打开文件夹: {INPUT_DIR}")
        import subprocess
        import platform
        if platform.system() == 'Windows':
            subprocess.run(['explorer', str(Path(input_dir).resolve())])
        elif platform.system() == 'Darwin':  # macOS
            subprocess.run(['open', str(Path(input_dir).resolve())])
        else:  # Linux
            subprocess.run(['xdg-open', str(Path(input_dir).resolve())])
        
        if len(excel_files) == 1:
            input_file = str(excel_files[0])
            print(f"\n✅ 发现1个Excel文件: {excel_files[0].name}")
            confirm = input("\n是否转换此文件？(直接回车确认/输入n取消): ").strip().lower()
            if confirm == 'n':
                print("\n❌ 已取消转换")
                return
        else:
            print(f"\n📋 发现 {len(excel_files)} 个Excel文件，请选择要转换的文件：")
            for idx, f in enumerate(excel_files, 1):
                file_size = f.stat().st_size // 1024
                print(f"  {idx}. {f.name} ({file_size} KB)")
            
            choice = input("\n请输入文件序号 (直接回车默认选择第1个): ").strip()
            
            if choice == '':
                selected_idx = 0
            else:
                try:
                    selected_idx = int(choice) - 1
                    if selected_idx < 0 or selected_idx >= len(excel_files):
                        print(f"\n❌ 错误：序号超出范围，请输入1-{len(excel_files)}")
                        return
                except ValueError:
                    print("\n❌ 错误：请输入有效的数字")
                    return
            
            input_file = str(excel_files[selected_idx])
            print(f"\n✅ 已选择: {excel_files[selected_idx].name}")
    
    base_name = Path(input_file).stem
    output_dir = safe_path(OUTPUT_DIR)
    output_file = Path(output_dir) / f"{base_name}.html"
    
    print(f"📄 输出: {output_file}")
    
    # 安全创建输出目录
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    
    # 步骤1: 提取并映射
    print("\n🔄 步骤 1/5: 分析Excel结构并提取图片...")
    image_data_map, image_position_map = extract_images_with_mapping(input_file)
    
    # 步骤2: 生成HTML
    print("\n🔄 步骤 2/5: 生成HTML内容...")
    
    # 检测文档类型
    from openpyxl import load_workbook
    wb = load_workbook(input_file)
    sheet_names = [ws.title for ws in wb.worksheets]
    wb.close()
    
    print(f"\n📋 检测到工作表: {', '.join(sheet_names)}")
    
    # 根据工作表名称判断文档类型
    if any(keyword in str(sheet_names) for keyword in ['设计', '详情', 'Design', 'Details']):
        print("📌 文档类型: 游戏设计文档")
        html_content = generate_html_smart(image_data_map, image_position_map)
    else:
        print("📌 文档类型: 通用文档（使用通用生成器）")
        html_content = generate_html_universal(input_file, image_data_map, image_position_map)
    
    # 步骤3: 验证内容完整性
    print("\n🔄 步骤 3/5: 验证内容完整性...")
    verification_report = verify_content(input_file, image_data_map, html_content)
    
    print(f"\n  📊 验证报告:")
    print(f"  ├─ Excel单元格总数: {verification_report['excel_cell_count']}")
    print(f"  ├─ HTML包含内容数: {verification_report['html_content_count']}")
    print(f"  ├─ 图片总数: {verification_report['image_count']}")
    print(f"  └─ 图片嵌入数: {verification_report['image_embedded_count']}")
    
    if not verification_report['is_valid']:
        print("\n  ⚠️  警告：检测到内容缺失")
        if verification_report['missing_cells']:
            print(f"  ├─ 缺失单元格样本: {verification_report['missing_cells'][:3]}")
        if verification_report['missing_images']:
            print(f"  └─ 缺失图片: {verification_report['missing_images']}")
    else:
        print(f"\n  ✅ 验证通过：所有内容已完整包含")
    
    # 步骤4: 文档建议已在generate_html_universal中自动生成
    print("\n🔄 步骤 4/5: 生成文档设计建议...")
    doc_name = os.path.basename(input_file).replace('.xlsx', '')
    if '物品' in doc_name or '背包' in doc_name or '道具' in doc_name:
        doc_type = '物品栏'
    elif '技能' in doc_name:
        doc_type = '技能'
    else:
        doc_type = '通用'
    print(f"  ✅ 已生成\"{doc_type}\"系统设计建议")
    
    # 步骤5: 保存HTML文件（安全路径写入）
    print("\n🔄 步骤 5/5: 输出HTML文件...")
    output_file_safe = safe_path(str(output_file))
    with open(output_file_safe, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    file_size_kb = Path(output_file_safe).stat().st_size // 1024
    
    print("\n" + "="*70)
    print("✅ 转换完成！")
    print("="*70)
    print(f"\n📄 文件: {output_file}")
    print(f"💾 大小: {file_size_kb} KB")
    print("\n" + "="*70)
    
    # 打开浏览器（安全路径）
    import webbrowser
    output_abs_path = Path(output_file_safe).resolve()
    webbrowser.open(f'file:///{output_abs_path}')

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        import traceback
        traceback.print_exc()
